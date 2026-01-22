from typing import List, Dict, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


class ExcelStyles:
    def __init__(self):
        self.header_font = Font(bold=True, color="FFFFFF", size=12)
        self.header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.center_alignment = Alignment(horizontal="center", vertical="center")
        self.left_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        self.currency_alignment = Alignment(horizontal="right", vertical="center")
        self.action_highlight = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
        self.price_highlight = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")


class SheetBuilder:
    CAPITAL_PER_OPPORTUNITY = 1500.0
    
    def __init__(self, worksheet: Worksheet, styles: ExcelStyles):
        self.ws = worksheet
        self.styles = styles
    
    def _sort_opportunities(self, opportunities: List[Dict]) -> List[Dict]:
        return sorted(
            opportunities,
            key=lambda x: x['arbitrage']['best_strategy']['roi_percent'] if x['arbitrage']['best_strategy'] else 0,
            reverse=True
        )
    
    def _build_headers(self, platform1_name: str, platform2_name: str, include_orderbook: bool = False, include_platform2_orderbook: bool = False) -> List[str]:
        headers = [
            "Rank",
            "Question",
            "ROI %",
        ]

        # Add Orderbook ROI columns only when orderbook data is available
        if include_orderbook or include_platform2_orderbook:
            headers.append("ROI ASK 1")
            headers.append("ROI ASK 2")

        headers.extend([
            "Bet YES on",
            "Bet NO on",
            "YES",
            "NO",
            "YES",
            "NO",
        ])

        if include_orderbook:
            # Platform1 orderbook columns (Polymarket)
            headers.extend([
                "Bid 1",
                "Bid 1 Size $",
                "Bid 2",
                "Bid 2 Size $",
                "Ask 1",
                "Ask 1 Size $",
                "Ask 2",
                "Ask 2 Size $",
                "Bid 1",
                "Bid 1 Size $",
                "Bid 2",
                "Bid 2 Size $",
                "Ask 1",
                "Ask 1 Size $",
                "Ask 2",
                "Ask 2 Size $",
            ])
        
        if include_platform2_orderbook:
            # Platform2 orderbook columns (predict.fun)
            headers.extend([
                "Bid 1",
                "Bid 1 Size $",
                "Bid 2",
                "Bid 2 Size $",
                "Ask 1",
                "Ask 1 Size $",
                "Ask 2",
                "Ask 2 Size $",
                "Bid 1",
                "Bid 1 Size $",
                "Bid 2",
                "Bid 2 Size $",
                "Ask 1",
                "Ask 1 Size $",
                "Ask 2",
                "Ask 2 Size $",
            ])
        
        return headers
    
    def _write_parent_headers(self, platform1_name: str, platform2_name: str, include_orderbook: bool = False, include_platform2_orderbook: bool = False, has_orderbook_roi: bool = False):
        col_offset = 2 if has_orderbook_roi else 0
        
        # Merge cells for Platform1 YES/NO parent header
        platform1_start = 6 + col_offset
        platform1_end = 7 + col_offset
        self.ws.merge_cells(start_row=1, start_column=platform1_start, end_row=1, end_column=platform1_end)
        cell = self.ws.cell(row=1, column=platform1_start)
        cell.value = platform1_name
        cell.font = self.styles.header_font
        cell.fill = self.styles.header_fill
        cell.alignment = self.styles.header_alignment
        cell.border = self.styles.border
        
        # Merge cells for Platform2 YES/NO parent header
        platform2_start = 8 + col_offset
        platform2_end = 9 + col_offset
        self.ws.merge_cells(start_row=1, start_column=platform2_start, end_row=1, end_column=platform2_end)
        cell = self.ws.cell(row=1, column=platform2_start)
        cell.value = platform2_name
        cell.font = self.styles.header_font
        cell.fill = self.styles.header_fill
        cell.alignment = self.styles.header_alignment
        cell.border = self.styles.border
        
        # Merge cells for Platform1 Orderbook YES parent header (8 columns)
        if include_orderbook:
            orderbook_start = 10 + col_offset
            yes_end = orderbook_start + 7
            self.ws.merge_cells(start_row=1, start_column=orderbook_start, end_row=1, end_column=yes_end)
            cell = self.ws.cell(row=1, column=orderbook_start)
            cell.value = f"{platform1_name} YES"
            cell.font = self.styles.header_font
            cell.fill = self.styles.header_fill
            cell.alignment = self.styles.header_alignment
            cell.border = self.styles.border
            
            # Merge cells for Platform1 Orderbook NO parent header
            no_start = yes_end + 1
            no_end = no_start + 7
            self.ws.merge_cells(start_row=1, start_column=no_start, end_row=1, end_column=no_end)
            cell = self.ws.cell(row=1, column=no_start)
            cell.value = f"{platform1_name} NO"
            cell.font = self.styles.header_font
            cell.fill = self.styles.header_fill
            cell.alignment = self.styles.header_alignment
            cell.border = self.styles.border
        
        # Merge cells for Platform2 Orderbook YES parent header (8 columns)
        if include_platform2_orderbook:
            # Start after platform1 orderbook columns (if present)
            if include_orderbook:
                orderbook2_start = 10 + col_offset + 16  # After platform1's 16 columns
            else:
                orderbook2_start = 10 + col_offset
            
            yes_end = orderbook2_start + 7
            self.ws.merge_cells(start_row=1, start_column=orderbook2_start, end_row=1, end_column=yes_end)
            cell = self.ws.cell(row=1, column=orderbook2_start)
            cell.value = f"{platform2_name} YES"
            cell.font = self.styles.header_font
            cell.fill = self.styles.header_fill
            cell.alignment = self.styles.header_alignment
            cell.border = self.styles.border
            
            # Merge cells for Platform2 Orderbook NO parent header
            no_start = yes_end + 1
            no_end = no_start + 7
            self.ws.merge_cells(start_row=1, start_column=no_start, end_row=1, end_column=no_end)
            cell = self.ws.cell(row=1, column=no_start)
            cell.value = f"{platform2_name} NO"
            cell.font = self.styles.header_font
            cell.fill = self.styles.header_fill
            cell.alignment = self.styles.header_alignment
            cell.border = self.styles.border
    
    def _write_headers(self, headers: List[str]):
        for col_num, header in enumerate(headers, 1):
            cell = self.ws.cell(row=2, column=col_num)
            cell.value = header
            cell.font = self.styles.header_font
            cell.fill = self.styles.header_fill
            cell.alignment = self.styles.header_alignment
            cell.border = self.styles.border
    
    def _calculate_shares(self, arb: Dict, best_strategy: Dict) -> tuple:
        strategy_type = best_strategy['type']
        
        if 'Yes on App1, No on App2' in strategy_type:
            price_app1 = arb['market1_yes']
            price_app2 = arb['market2_no']
        else:
            price_app1 = arb['market1_no']
            price_app2 = arb['market2_yes']
        
        total_cost = price_app1 + price_app2
        
        bet_app1 = self.CAPITAL_PER_OPPORTUNITY * (price_app1 / total_cost) if total_cost > 0 else 0
        bet_app2 = self.CAPITAL_PER_OPPORTUNITY * (price_app2 / total_cost) if total_cost > 0 else 0
        
        shares_app1 = bet_app1 / price_app1 if price_app1 > 0 else 0
        shares_app2 = bet_app2 / price_app2 if price_app2 > 0 else 0
        
        return shares_app1, shares_app2
    
    def _write_cell(self, row: int, col: int, value, alignment, number_format: str = None, fill=None):
        cell = self.ws.cell(row=row, column=col, value=value)
        cell.alignment = alignment
        cell.border = self.styles.border
        if number_format:
            cell.number_format = number_format
        if fill:
            cell.fill = fill
    
    def _write_basic_columns(self, row: int, idx: int, market: Dict, best: Dict, arb: Dict, has_orderbook_roi: bool = False):
        actual_row = row + 1
        self._write_cell(actual_row, 1, idx, self.styles.center_alignment)
        self._write_cell(actual_row, 2, market['question'], self.styles.left_alignment)
        self._write_cell(actual_row, 3, round(best['roi_percent'], 2), self.styles.center_alignment, '0.00"%"')
    
    def _write_orderbook_roi_columns(self, row: int, opp: Dict):
        actual_row = row + 1
        orderbook_roi = opp.get('orderbook_roi_percent')
        if orderbook_roi is not None:
            self._write_cell(actual_row, 4, round(orderbook_roi, 2), self.styles.center_alignment, '0.00"%"')

        orderbook_roi_ask2 = opp.get('orderbook_roi_ask2_percent')
        if orderbook_roi_ask2 is not None:
            self._write_cell(actual_row, 5, round(orderbook_roi_ask2, 2), self.styles.center_alignment, '0.00"%"')

    def _write_price_columns(self, row: int, arb: Dict, best_strategy: Dict, platform1_name: str, platform2_name: str, col_offset: int = 0):
        actual_row = row + 1
        strategy_type = best_strategy['type']
        
        # Determine which prices to highlight based on strategy
        if 'Yes on App1, No on App2' in strategy_type:
            # Highlight Platform1 YES and Platform2 NO
            self._write_cell(actual_row, 6 + col_offset, round(arb['market1_yes'], 3), self.styles.currency_alignment, '$0.000', self.styles.price_highlight)
            self._write_cell(actual_row, 7 + col_offset, round(arb['market1_no'], 3), self.styles.currency_alignment, '$0.000')
            self._write_cell(actual_row, 8 + col_offset, round(arb['market2_yes'], 3), self.styles.currency_alignment, '$0.000')
            self._write_cell(actual_row, 9 + col_offset, round(arb['market2_no'], 3), self.styles.currency_alignment, '$0.000', self.styles.price_highlight)
        else:
            # Highlight Platform1 NO and Platform2 YES
            self._write_cell(actual_row, 6 + col_offset, round(arb['market1_yes'], 3), self.styles.currency_alignment, '$0.000')
            self._write_cell(actual_row, 7 + col_offset, round(arb['market1_no'], 3), self.styles.currency_alignment, '$0.000', self.styles.price_highlight)
            self._write_cell(actual_row, 8 + col_offset, round(arb['market2_yes'], 3), self.styles.currency_alignment, '$0.000', self.styles.price_highlight)
            self._write_cell(actual_row, 9 + col_offset, round(arb['market2_no'], 3), self.styles.currency_alignment, '$0.000')

    def _write_orderbook_columns(self, row: int, orderbook_data: Dict, best_strategy: Dict, col_offset: int = 0):
        if not orderbook_data:
            return
        
        actual_row = row + 1
        strategy_type = best_strategy.get('type', '')
        highlight_yes = 'Yes on App1' in strategy_type
        highlight_no = 'No on App1' in strategy_type
        
        col = 10 + col_offset
        orderbook_fields = [
            ('yes_bid1_price', '$0.000', False),
            ('yes_bid1_size_usd', '$0.00', False),  # Bid sizes never highlighted
            ('yes_bid2_price', '$0.000', False),
            ('yes_bid2_size_usd', '$0.00', False),  # Bid sizes never highlighted
            ('yes_ask1_price', '$0.000', False),
            ('yes_ask1_size_usd', '$0.00', highlight_yes),  # Highlight YES Ask sizes when buying YES
            ('yes_ask2_price', '$0.000', False),
            ('yes_ask2_size_usd', '$0.00', highlight_yes),  # Highlight YES Ask sizes when buying YES
            ('no_bid1_price', '$0.000', False),
            ('no_bid1_size_usd', '$0.00', False),  # Bid sizes never highlighted
            ('no_bid2_price', '$0.000', False),
            ('no_bid2_size_usd', '$0.00', False),  # Bid sizes never highlighted
            ('no_ask1_price', '$0.000', False),
            ('no_ask1_size_usd', '$0.00', highlight_no),  # Highlight NO Ask sizes when buying NO
            ('no_ask2_price', '$0.000', False),
            ('no_ask2_size_usd', '$0.00', highlight_no),  # Highlight NO Ask sizes when buying NO
        ]
        
        for field, num_format, should_highlight in orderbook_fields:
            value = orderbook_data.get(field)
            if value is not None:
                precision = 3 if 'price' in field else 2
                fill = self.styles.action_highlight if should_highlight else None
                self._write_cell(actual_row, col, round(value, precision), self.styles.currency_alignment, num_format, fill)
            col += 1
    
    def _write_platform2_orderbook_columns(self, row: int, orderbook_data: Dict, best_strategy: Dict, col_offset: int = 0, platform1_has_orderbook: bool = False):
        if not orderbook_data:
            return
        
        actual_row = row + 1
        strategy_type = best_strategy.get('type', '')
        # For platform2, check if strategy is buying YES/NO on App2
        highlight_yes = 'Yes on App2' in strategy_type
        highlight_no = 'No on App2' in strategy_type
        
        # Calculate starting column: after platform1 orderbook if it exists
        base_col = 10 + col_offset
        if platform1_has_orderbook:
            base_col += 16  # Platform1 orderbook has 16 columns
        
        col = base_col
        orderbook_fields = [
            ('yes_bid1_price', '$0.000', False),
            ('yes_bid1_size_usd', '$0.00', False),
            ('yes_bid2_price', '$0.000', False),
            ('yes_bid2_size_usd', '$0.00', False),
            ('yes_ask1_price', '$0.000', False),
            ('yes_ask1_size_usd', '$0.00', highlight_yes),  # Highlight YES Ask sizes when buying YES on App2
            ('yes_ask2_price', '$0.000', False),
            ('yes_ask2_size_usd', '$0.00', highlight_yes),
            ('no_bid1_price', '$0.000', False),
            ('no_bid1_size_usd', '$0.00', False),
            ('no_bid2_price', '$0.000', False),
            ('no_bid2_size_usd', '$0.00', False),
            ('no_ask1_price', '$0.000', False),
            ('no_ask1_size_usd', '$0.00', highlight_no),  # Highlight NO Ask sizes when buying NO on App2
            ('no_ask2_price', '$0.000', False),
            ('no_ask2_size_usd', '$0.00', highlight_no),
        ]
        
        for field, num_format, should_highlight in orderbook_fields:
            value = orderbook_data.get(field)
            if value is not None:
                precision = 3 if 'price' in field else 2
                fill = self.styles.action_highlight if should_highlight else None
                self._write_cell(actual_row, col, round(value, precision), self.styles.currency_alignment, num_format, fill)
            col += 1
    
    def _write_strategy_columns(self, row: int, best_strategy: Dict, platform1_name: str, platform2_name: str, col_offset: int = 0):
        actual_row = row + 1
        strategy_type = best_strategy['type']
        
        if 'Yes on App1, No on App2' in strategy_type:
            yes_platform = platform1_name
            no_platform = platform2_name
        else:
            yes_platform = platform2_name
            no_platform = platform1_name
        
        self._write_cell(actual_row, 4 + col_offset, yes_platform, self.styles.center_alignment)
        self._write_cell(actual_row, 5 + col_offset, no_platform, self.styles.center_alignment)

    def _set_column_widths(self, column_widths: Dict[int, int]):
        for col_num, width in column_widths.items():
            self.ws.column_dimensions[get_column_letter(col_num)].width = width
    
    def _finalize_sheet(self):
        self.ws.freeze_panes = "A3"
        self.ws.auto_filter.ref = self.ws.dimensions
    
    def build_sheet(self, opportunities: List[Dict],
                   platform1_name: str, platform2_name: str, 
                   platform1_widths: tuple = (10, 10), platform2_widths: tuple = (10, 10)):
        sorted_opportunities = self._sort_opportunities(opportunities)
        
        # Check if any opportunity has orderbook data
        has_orderbook = any(opp.get('polymarket_orderbook') for opp in sorted_opportunities)
        
        # Check for platform2 orderbook based on platform name
        has_platform2_orderbook = False
        if platform2_name == "predict.fun":
            has_platform2_orderbook = any(opp.get('predict_orderbook') for opp in sorted_opportunities)
        elif platform2_name == "Opinion":
            has_platform2_orderbook = any(opp.get('opinion_orderbook') for opp in sorted_opportunities)
        
        # Check if any opportunity has orderbook ROI
        has_orderbook_roi = any(opp.get('orderbook_roi_percent') is not None for opp in sorted_opportunities)

        # Calculate column offset: +2 if orderbook ROI columns are present (ROI ASK 1 and ROI ASK 2)
        col_offset = 2 if has_orderbook_roi else 0

        headers = self._build_headers(platform1_name, platform2_name, include_orderbook=has_orderbook, include_platform2_orderbook=has_platform2_orderbook)
        
        # Write parent headers first (row 1) with merged cells
        self._write_parent_headers(platform1_name, platform2_name, include_orderbook=has_orderbook, include_platform2_orderbook=has_platform2_orderbook, has_orderbook_roi=has_orderbook_roi)
        
        # Write sub-headers (row 2) and add header styling to non-merged cells in row 1
        self._write_headers(headers)
        for col_num in range(1, 6 + col_offset):
            cell = self.ws.cell(row=1, column=col_num)
            if cell.value is None:
                cell.value = ""
            cell.font = self.styles.header_font
            cell.fill = self.styles.header_fill
            cell.alignment = self.styles.header_alignment
            cell.border = self.styles.border
        
        for idx, opp in enumerate(sorted_opportunities, 1):
            market = opp['market']
            arb = opp['arbitrage']
            best = arb['best_strategy']
            
            if not best:
                continue
            
            row = idx + 1
            
            self._write_basic_columns(row, idx, market, best, arb, has_orderbook_roi)

            # Write orderbook ROI columns if available
            if has_orderbook_roi:
                self._write_orderbook_roi_columns(row, opp)

            self._write_strategy_columns(row, best, platform1_name, platform2_name, col_offset)
            self._write_price_columns(row, arb, best, platform1_name, platform2_name, col_offset)

            # Write platform1 orderbook data if available
            orderbook_data = opp.get('polymarket_orderbook')
            if orderbook_data:
                self._write_orderbook_columns(row, orderbook_data, best, col_offset)
            
            # Write platform2 orderbook data if available based on platform name
            platform2_orderbook = None
            if platform2_name == "predict.fun":
                platform2_orderbook = opp.get('predict_orderbook')
            elif platform2_name == "Opinion":
                platform2_orderbook = opp.get('opinion_orderbook')
            
            if platform2_orderbook:
                self._write_platform2_orderbook_columns(row, platform2_orderbook, best, col_offset, platform1_has_orderbook=has_orderbook)

        column_widths = {
            1: 6,
            2: 50,
            3: 10,
        }

        # Add orderbook ROI column widths if present
        if has_orderbook_roi:
            column_widths[4] = 12  # ROI ASK 1
            column_widths[5] = 12  # ROI ASK 2
            # Shift all subsequent columns by 2
            column_widths.update({
                6: 13,  # Bet YES on
                7: 13,  # Bet NO on
                8: 12,  # Platform1 YES
                9: 12,  # Platform1 NO
                10: 12,  # Platform2 YES
                11: 12,  # Platform2 NO
            })
        else:
            # No orderbook ROI columns
            column_widths.update({
                4: 13,  # Bet YES on
                5: 13,  # Bet NO on
                6: 12,  # Platform1 YES
                7: 12,  # Platform1 NO
                8: 12,  # Platform2 YES
                9: 12,  # Platform2 NO
            })

        # Add platform1 orderbook column widths if present
        if has_orderbook:
            orderbook_start_col = 10 + col_offset
            column_widths.update({
                orderbook_start_col: 11,  # YES Bid 1
                orderbook_start_col + 1: 14,  # YES Bid 1 Size $
                orderbook_start_col + 2: 11,  # YES Bid 2
                orderbook_start_col + 3: 14,  # YES Bid 2 Size $
                orderbook_start_col + 4: 11,  # YES Ask 1
                orderbook_start_col + 5: 14,  # YES Ask 1 Size $
                orderbook_start_col + 6: 11,  # YES Ask 2
                orderbook_start_col + 7: 14,  # YES Ask 2 Size $
                orderbook_start_col + 8: 11,  # NO Bid 1
                orderbook_start_col + 9: 14,  # NO Bid 1 Size $
                orderbook_start_col + 10: 11,  # NO Bid 2
                orderbook_start_col + 11: 14,  # NO Bid 2 Size $
                orderbook_start_col + 12: 11,  # NO Ask 1
                orderbook_start_col + 13: 14,  # NO Ask 1 Size $
                orderbook_start_col + 14: 11,  # NO Ask 2
                orderbook_start_col + 15: 14,  # NO Ask 2 Size $
            })
        
        # Add platform2 orderbook column widths if present
        if has_platform2_orderbook:
            # Start after platform1 orderbook columns (if present)
            if has_orderbook:
                orderbook2_start_col = 10 + col_offset + 16
            else:
                orderbook2_start_col = 10 + col_offset
            
            column_widths.update({
                orderbook2_start_col: 11,  # YES Bid 1
                orderbook2_start_col + 1: 14,  # YES Bid 1 Size $
                orderbook2_start_col + 2: 11,  # YES Bid 2
                orderbook2_start_col + 3: 14,  # YES Bid 2 Size $
                orderbook2_start_col + 4: 11,  # YES Ask 1
                orderbook2_start_col + 5: 14,  # YES Ask 1 Size $
                orderbook2_start_col + 6: 11,  # YES Ask 2
                orderbook2_start_col + 7: 14,  # YES Ask 2 Size $
                orderbook2_start_col + 8: 11,  # NO Bid 1
                orderbook2_start_col + 9: 14,  # NO Bid 1 Size $
                orderbook2_start_col + 10: 11,  # NO Bid 2
                orderbook2_start_col + 11: 14,  # NO Bid 2 Size $
                orderbook2_start_col + 12: 11,  # NO Ask 1
                orderbook2_start_col + 13: 14,  # NO Ask 1 Size $
                orderbook2_start_col + 14: 11,  # NO Ask 2
                orderbook2_start_col + 15: 14,  # NO Ask 2 Size $
            })
        
        self._set_column_widths(column_widths)
        self._finalize_sheet()


def generate_excel_report(opportunities: List[Dict], output_path: str, 
                         opinion_opportunities: Optional[List[Dict]] = None,
                         opinion_vs_predict_opportunities: Optional[List[Dict]] = None):
    wb = Workbook()
    styles = ExcelStyles()
    
    ws = wb.active
    ws.title = "Polymarket vs predict.fun"
    
    builder = SheetBuilder(ws, styles)
    builder.build_sheet(opportunities, "Polymarket", "predict.fun", 
                       platform1_widths=(15, 17), platform2_widths=(15, 17))
    
    # Always create Opinion sheet if opinion_opportunities is provided (even if empty)
    if opinion_opportunities is not None:
        ws_opinion = wb.create_sheet(title="Polymarket vs Opinion")
        builder_opinion = SheetBuilder(ws_opinion, styles)
        if len(opinion_opportunities) > 0:
            builder_opinion.build_sheet(opinion_opportunities, 
                                       "Polymarket", "Opinion",
                                       platform1_widths=(15, 17), platform2_widths=(12, 14))
        else:
            # Create empty sheet with headers only
            builder_opinion.build_sheet([], "Polymarket", "Opinion",
                                       platform1_widths=(15, 17), platform2_widths=(12, 14))
    
    # Create Opinion vs predict.fun sheet
    if opinion_vs_predict_opportunities is not None:
        ws_opinion_predict = wb.create_sheet(title="Opinion vs predict.fun")
        builder_opinion_predict = SheetBuilder(ws_opinion_predict, styles)
        if len(opinion_vs_predict_opportunities) > 0:
            builder_opinion_predict.build_sheet(opinion_vs_predict_opportunities,
                                               "Opinion", "predict.fun",
                                               platform1_widths=(12, 14), platform2_widths=(15, 17))
        else:
            # Create empty sheet with headers only
            builder_opinion_predict.build_sheet([], "Opinion", "predict.fun",
                                               platform1_widths=(12, 14), platform2_widths=(15, 17))
    
    wb.save(output_path)
